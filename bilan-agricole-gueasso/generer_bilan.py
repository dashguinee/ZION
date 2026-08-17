#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genere le classeur comptable des campagnes agricoles 2024 et 2025
Domaine de 60 ha - Sous-prefecture de Gueasso, Prefecture de Lola (Guinee forestiere)
Exploitant : KASI GROUP SARL (RCCM GN.TCC.2025.B.16465 - NIF 642333652)

Referentiel : plan comptable SYSCOHADA revise.
Nature du document : comptes d'exploitation et bilans RECONSTITUES sur hypotheses
technico-economiques explicites (voir onglet "Notice"). Ce ne sont pas des etats
financiers certifies.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- styles
POLICE = "Arial"

F_TITRE = Font(name=POLICE, size=16, bold=True, color="1F3864")
F_SOUSTITRE = Font(name=POLICE, size=10, italic=True, color="404040")
F_SECTION = Font(name=POLICE, size=11, bold=True, color="FFFFFF")
F_ENTETE = Font(name=POLICE, size=10, bold=True, color="FFFFFF")
F_NORMAL = Font(name=POLICE, size=10)
F_GRAS = Font(name=POLICE, size=10, bold=True)
F_INPUT = Font(name=POLICE, size=10, color="0000FF")
F_LIEN = Font(name=POLICE, size=10, color="008000")
F_TOTAL = Font(name=POLICE, size=10, bold=True, color="1F3864")
F_NOTE = Font(name=POLICE, size=9, italic=True, color="595959")

FILL_SECTION = PatternFill("solid", fgColor="1F3864")
FILL_ENTETE = PatternFill("solid", fgColor="2E75B6")
FILL_TOTAL = PatternFill("solid", fgColor="D9E2F3")
FILL_CLE = PatternFill("solid", fgColor="FFFF00")
FILL_ALT = PatternFill("solid", fgColor="F2F2F2")

TRAIT = Side(style="thin", color="BFBFBF")
BORDURE = Border(left=TRAIT, right=TRAIT, top=TRAIT, bottom=TRAIT)

FMT_GNF = '#,##0;(#,##0);-'
FMT_T = '#,##0.00;(#,##0.00);-'
FMT_PCT = '0.0%'
FMT_HA = '#,##0.0;(#,##0.0);-'
FMT_AN = '0.00'

wb = Workbook()


def sec(ws, row, texte, ncols=6):
    """Bandeau de section."""
    ws.cell(row=row, column=1, value=texte)
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = FILL_SECTION
        ws.cell(row=row, column=c).font = F_SECTION
    ws.row_dimensions[row].height = 20


def entete(ws, row, libelles, col0=1):
    for i, lib in enumerate(libelles):
        c = ws.cell(row=row, column=col0 + i, value=lib)
        c.fill = FILL_ENTETE
        c.font = F_ENTETE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDURE
    ws.row_dimensions[row].height = 30


def ligne(ws, row, libelle, valeurs, fmt=FMT_GNF, font=F_NORMAL,
          fill=None, col0=2, indent=0):
    c = ws.cell(row=row, column=1, value=libelle)
    c.font = font
    c.alignment = Alignment(indent=indent)
    c.border = BORDURE
    if fill:
        c.fill = fill
    for i, v in enumerate(valeurs):
        cc = ws.cell(row=row, column=col0 + i, value=v)
        cc.number_format = fmt
        cc.font = font
        cc.border = BORDURE
        if fill:
            cc.fill = fill
    return row


def titre_page(ws, titre, sous_titre, largeur=6):
    ws.cell(row=1, column=1, value=titre).font = F_TITRE
    ws.cell(row=2, column=1, value=sous_titre).font = F_SOUSTITRE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=largeur)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=largeur)
    ws.row_dimensions[1].height = 24


# =====================================================================
# 1. NOTICE
# =====================================================================
ws = wb.active
ws.title = "Notice"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 118

lignes_notice = [
    ("T", "BILAN COMPTABLE DES CAMPAGNES AGRICOLES 2024 ET 2025"),
    ("S", "Domaine agricole de 60 hectares - Sous-prefecture de Gueasso, Prefecture de Lola, Region de N'Zerekore"),
    ("", ""),
    ("H", "1. IDENTIFICATION DE L'EXPLOITANT"),
    ("N", "Raison sociale        : KASI GROUP SARL (sigle K-G-SARL)"),
    ("N", "Forme juridique       : Societe a responsabilite limitee"),
    ("N", "RCCM                  : GN.TCC.2025.B.16465 - Formalite GN.TCC.2025.18612"),
    ("N", "NIF                   : 642333652 - Numero eTax : 1228802510892"),
    ("N", "Date d'immatriculation: 19 novembre 2025, Tribunal de Commerce de Conakry"),
    ("N", "Capital social        : 10 000 000 GNF integralement libere en numeraire"),
    ("N", "Siege social          : Immeuble Sonoco, Marche Niger, Commune de Kaloum, Conakry"),
    ("N", "Cogerants             : KABA Alhassane et SIDIBE Mohamed"),
    ("N", "Objet social          : BTP, commerce, AGRICULTURE, communication, technologie, transport, mines, elevage, immobilier"),
    ("", ""),
    ("H", "2. NATURE ET PORTEE DU DOCUMENT - A LIRE IMPERATIVEMENT"),
    ("A", "KASI GROUP SARL n'est immatriculee que depuis le 19 novembre 2025. La societe ne peut donc pas"),
    ("A", "produire d'etats financiers historiques certifies au titre de la campagne 2024, ni d'un exercice social"),
    ("A", "complet 2025. Le present document est un jeu de COMPTES D'EXPLOITATION ET DE BILANS RECONSTITUES,"),
    ("A", "etabli sur la base d'hypotheses technico-economiques explicites, entierement parametrees dans l'onglet"),
    ("A", "\"Hypotheses\". Il n'est ni audite, ni certifie, ni opposable a l'administration fiscale."),
    ("", ""),
    ("N", "Cadre retenu pour chacune des deux campagnes :"),
    ("N", "  - Campagne 2024 : exploitation conduite en nom propre par les promoteurs (regime de l'entreprise"),
    ("N", "    individuelle agricole), avant constitution de la societe. Comptes tenus en partie double."),
    ("N", "  - Campagne 2025 : exploitation apportee a KASI GROUP SARL a sa constitution. Les capitaux investis"),
    ("N", "    par les promoteurs en 2024 sont reclasses en comptes courants d'associes bloques au 31/12/2025."),
    ("", ""),
    ("N", "Usage recommande : dossier de financement, demande de credit agricole, presentation a un partenaire"),
    ("N", "technique ou financier, plan d'affaires. Toute production devant un tiers doit conserver le present"),
    ("N", "avertissement et la mention \"comptes reconstitues sur hypotheses\"."),
    ("", ""),
    ("H", "3. REFERENTIEL ET CONVENTIONS"),
    ("N", "Referentiel comptable  : SYSCOHADA revise (Acte uniforme OHADA relatif au droit comptable, systeme normal)"),
    ("N", "Monnaie de tenue       : franc guineen (GNF). Tous les montants sont exprimes en GNF."),
    ("N", "Cout historique        : les immobilisations sont evaluees au cout d'acquisition."),
    ("N", "Amortissements         : mode lineaire, calcul prorata temporis a compter de la mise en service."),
    ("N", "Stocks                 : intrants au cout d'achat, produits finis au cout de production."),
    ("N", "Imposition             : impot sur les benefices au taux de 25 %, hors regimes derogatoires agricoles."),
    ("N", "Exercice               : la campagne agricole est assimilee a l'exercice civil (1er janvier - 31 decembre)."),
    ("", ""),
    ("H", "4. CONTEXTE AGRO-ECOLOGIQUE DE LA ZONE"),
    ("N", "Gueasso, prefecture de Lola, se situe en Guinee forestiere. Pluviometrie annuelle de 1 800 a 2 200 mm"),
    ("N", "repartie sur 8 a 9 mois, sols ferrallitiques desatures profonds, relief de collines et de bas-fonds."),
    ("N", "Cette zone est favorable au riz pluvial et de bas-fond, au mais, au manioc et a l'arachide, qui"),
    ("N", "constituent l'assolement retenu. Debouches : marches de Lola, N'Zerekore, Beyla et flux transfrontaliers."),
    ("", ""),
    ("H", "5. LOGIQUE DE LA PROGRESSION ENTRE LES DEUX CAMPAGNES"),
    ("N", "2024 - campagne de mise en valeur. Defriche et amenagement du domaine, mecanisation confiee a des"),
    ("N", "prestataires, transformation (decorticage, gari) sous-traitee, pertes post-recolte de 10 %."),
    ("N", "Resultat positif mais etroit : la campagne absorbe les couts de premiere installation."),
    ("N", "2025 - campagne de consolidation. Acquisition d'un tracteur et d'une unite post-recolte, amenagement"),
    ("N", "de 4 ha de bas-fond rizicole, semences certifiees, fumure raisonnee sur analyse de sol, encadrement"),
    ("N", "technique, pertes ramenees a 5 % et integration de la transformation qui capte la marge d'usinage."),
    ("N", "Le resultat net est multiplie par pres de 3,7 et la valeur ajoutee a l'hectare par plus de 2."),
    ("", ""),
    ("H", "6. CODE COULEUR DU CLASSEUR"),
    ("N", "Texte bleu   : donnee d'entree modifiable (hypothese, parametre, montant saisi)"),
    ("N", "Texte noir   : resultat calcule par formule"),
    ("N", "Texte vert   : lien vers un autre onglet"),
    ("N", "Fond jaune   : hypothese cle ayant un effet majeur sur le resultat"),
    ("", ""),
    ("H", "7. ORGANISATION DU CLASSEUR"),
    ("N", "Hypotheses        : tous les parametres techniques, commerciaux et financiers"),
    ("N", "Production        : tonnages produits, affectation entre vente brute et transformation, chiffre d'affaires"),
    ("N", "Charges           : charges d'exploitation detaillees par compte SYSCOHADA"),
    ("N", "Immobilisations   : plan d'investissement et tableau d'amortissement"),
    ("N", "Compte de resultat: compte de resultat comparatif 2024 / 2025"),
    ("N", "Bilan             : bilan comparatif au 31/12/2024 et au 31/12/2025, avec controle d'equilibre"),
    ("N", "Flux & Financement: tableau des flux de tresorerie et plan de financement"),
    ("N", "SIG & Ratios      : soldes intermediaires de gestion, ratios, seuil de rentabilite"),
    ("N", "Comparatif        : synthese de la progression 2024 vers 2025"),
]

r = 1
for typ, txt in lignes_notice:
    c = ws.cell(row=r, column=1, value=txt)
    if typ == "T":
        c.font = F_TITRE
        ws.row_dimensions[r].height = 24
    elif typ == "S":
        c.font = F_SOUSTITRE
    elif typ == "H":
        c.font = Font(name=POLICE, size=11, bold=True, color="FFFFFF")
        c.fill = FILL_SECTION
        ws.row_dimensions[r].height = 18
    elif typ == "A":
        c.font = Font(name=POLICE, size=10, bold=True, color="C00000")
    else:
        c.font = F_NORMAL
    r += 1


# =====================================================================
# 2. HYPOTHESES
# =====================================================================
hy = wb.create_sheet("Hypotheses")
hy.sheet_view.showGridLines = False
hy.column_dimensions["A"].width = 58
for col in "BCD":
    hy.column_dimensions[col].width = 18
hy.column_dimensions["E"].width = 62

titre_page(hy, "HYPOTHESES TECHNICO-ECONOMIQUES",
           "Toutes les cellules bleues sont modifiables : le classeur entier se recalcule.", 5)

sec(hy, 4, "A. IDENTIFICATION ET PERIMETRE", 5)
ident = [
    ("Raison sociale de l'exploitant", "KASI GROUP SARL"),
    ("RCCM", "GN.TCC.2025.B.16465"),
    ("NIF", "642333652"),
    ("Localisation du domaine", "Gueasso, Prefecture de Lola, Region de N'Zerekore"),
    ("Superficie totale du domaine (ha)", 60),
    ("Monnaie de tenue des comptes", "GNF"),
]
r = 5
for lib, val in ident:
    hy.cell(row=r, column=1, value=lib).font = F_GRAS
    c = hy.cell(row=r, column=2, value=val)
    c.font = F_INPUT
    if isinstance(val, (int, float)):
        c.number_format = FMT_HA
    hy.cell(row=r, column=1).border = BORDURE
    c.border = BORDURE
    r += 1
LIG_SUPERFICIE = 9  # B9 = 60 ha

sec(hy, 12, "B. ASSOLEMENT (hectares emblaves)", 5)
entete(hy, 13, ["Speculation", "Campagne 2024", "Campagne 2025", "Variation", "Justification"])
assolement = [
    ("Riz pluvial NERICA", 22, 23, "Speculation pivot, debouche local et regional assure"),
    ("Riz de bas-fond amenage", 0, 4, "4 ha de bas-fond planes et endigues en 2025, rendement double"),
    ("Mais grain", 14, 14, "Rotation et debouche provende / brasserie"),
    ("Manioc", 12, 11, "Surface reduite au profit du riz, plus rentable apres integration de l'usinage"),
    ("Arachide (coque)", 7, 6, "Legumineuse de rotation, fixation azotee, tete d'assolement"),
]
r = 14
for lib, a24, a25, just in assolement:
    hy.cell(row=r, column=1, value=lib).font = F_NORMAL
    for i, v in enumerate([a24, a25]):
        c = hy.cell(row=r, column=2 + i, value=v)
        c.font = F_INPUT
        c.number_format = FMT_HA
    hy.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = FMT_HA
    hy.cell(row=r, column=4).font = F_NORMAL
    hy.cell(row=r, column=5, value=just).font = F_NOTE
    for c in range(1, 6):
        hy.cell(row=r, column=c).border = BORDURE
    r += 1
LIG_ASSOL_DEB, LIG_ASSOL_FIN = 14, 18
ligne(hy, 19, "Total emblave (ha)",
      [f"=SUM(B{LIG_ASSOL_DEB}:B{LIG_ASSOL_FIN})", f"=SUM(C{LIG_ASSOL_DEB}:C{LIG_ASSOL_FIN})",
       "=C19-B19"], FMT_HA, F_TOTAL, FILL_TOTAL)
ligne(hy, 20, "Superficie non cultivee (siege, hangars, pistes, pare-feu)",
      [f"=$B${LIG_SUPERFICIE}-B19", f"=$B${LIG_SUPERFICIE}-C19", "=C20-B20"], FMT_HA)
hy.cell(row=20, column=5, value="Emprise reduite en 2025 grace au reamenagement des pistes").font = F_NOTE
LIG_EMBLAVE = 19

sec(hy, 22, "C. RENDEMENTS (tonnes / hectare)", 5)
entete(hy, 23, ["Speculation", "Campagne 2024", "Campagne 2025", "Progression", "Justification agronomique"])
rendements = [
    ("Riz pluvial NERICA", 2.60, 3.30, "Semence certifiee R1, fumure fractionnee, desherbage precoce"),
    ("Riz de bas-fond amenage", 0.00, 4.50, "Maitrise de l'eau, repiquage en ligne, deux sarclages"),
    ("Mais grain", 2.90, 3.70, "Variete composite amelioree, densite 53 000 pieds/ha"),
    ("Manioc (racines fraiches)", 13.50, 17.00, "Boutures saines selectionnees, billonnage mecanise"),
    ("Arachide (gousses en coque)", 1.35, 1.70, "Semence renouvelee, traitement de semence, recolte a maturite"),
]
r = 24
for lib, y24, y25, just in rendements:
    hy.cell(row=r, column=1, value=lib).font = F_NORMAL
    for i, v in enumerate([y24, y25]):
        c = hy.cell(row=r, column=2 + i, value=v)
        c.font = F_INPUT
        c.number_format = FMT_T
        c.fill = FILL_CLE
    hy.cell(row=r, column=4, value=f'=IF(B{r}=0,"n.a.",C{r}/B{r}-1)').number_format = FMT_PCT
    hy.cell(row=r, column=4).font = F_NORMAL
    hy.cell(row=r, column=5, value=just).font = F_NOTE
    for c in range(1, 6):
        hy.cell(row=r, column=c).border = BORDURE
    r += 1
LIG_RDT_DEB = 24

sec(hy, 30, "D. PARAMETRES TECHNIQUES DE RECOLTE ET DE TRANSFORMATION", 5)
entete(hy, 31, ["Parametre", "Campagne 2024", "Campagne 2025", "", "Commentaire"])
techniques = [
    ("Taux de pertes post-recolte", 0.10, 0.05, FMT_PCT, True,
     "2025 : magasin ventile, sechage sur aire betonnee, palettisation"),
    ("Rendement a l'usinage du paddy (riz blanc / paddy)", 0.65, 0.65, FMT_PCT, False,
     "Norme technique d'une decortiqueuse a rouleaux caoutchouc"),
    ("Rendement de transformation manioc en gari", 0.22, 0.22, FMT_PCT, False,
     "Norme technique rapeuse-presse (1 t de racines donne 220 kg de gari)"),
    ("Part du paddy usinee en riz blanc", 0.55, 0.85, FMT_PCT, True,
     "2024 : prestation externe limitee. 2025 : unite propre de 1 t/h"),
    ("Part du manioc transformee en gari", 0.25, 0.45, FMT_PCT, True,
     "2025 : rapeuse-presse propre, marge d'usinage internalisee"),
]
r = 32
for lib, v24, v25, fmt, cle, com in techniques:
    hy.cell(row=r, column=1, value=lib).font = F_NORMAL
    for i, v in enumerate([v24, v25]):
        c = hy.cell(row=r, column=2 + i, value=v)
        c.font = F_INPUT
        c.number_format = fmt
        if cle:
            c.fill = FILL_CLE
    hy.cell(row=r, column=5, value=com).font = F_NOTE
    for c in range(1, 6):
        hy.cell(row=r, column=c).border = BORDURE
    r += 1
LIG_PERTES, LIG_USINAGE, LIG_GARI_TX, LIG_PART_USINE, LIG_PART_GARI = 32, 33, 34, 35, 36

sec(hy, 38, "E. PRIX DE VENTE MOYENS (GNF / kg)", 5)
entete(hy, 39, ["Produit", "Campagne 2024", "Campagne 2025", "Variation", "Reference de marche"])
prix = [
    ("Paddy (riz non usine), bord champ", 5000, 5400, "Marche de collecte de Lola et N'Zerekore"),
    ("Riz blanc local, sortie magasin", 11500, 12500, "Prix de gros N'Zerekore, sac de 50 kg"),
    ("Mais grain sec", 4200, 4600, "Demande provende et brasserie, prix de campagne"),
    ("Manioc frais (racines)", 1700, 1850, "Prix bord champ, vente en vrac aux collecteurs"),
    ("Gari (semoule de manioc)", 9000, 10000, "Prix de detail regional, conditionne en sacs de 25 kg"),
    ("Arachide en coque", 9500, 10500, "Prix de campagne, vente aux huileries et grossistes"),
]
r = 40
for lib, p24, p25, ref in prix:
    hy.cell(row=r, column=1, value=lib).font = F_NORMAL
    for i, v in enumerate([p24, p25]):
        c = hy.cell(row=r, column=2 + i, value=v)
        c.font = F_INPUT
        c.number_format = FMT_GNF
        c.fill = FILL_CLE
    hy.cell(row=r, column=4, value=f"=C{r}/B{r}-1").number_format = FMT_PCT
    hy.cell(row=r, column=4).font = F_NORMAL
    hy.cell(row=r, column=5, value=ref).font = F_NOTE
    for c in range(1, 6):
        hy.cell(row=r, column=c).border = BORDURE
    r += 1
LIG_PX_PADDY, LIG_PX_RIZ, LIG_PX_MAIS, LIG_PX_MANIOC, LIG_PX_GARI, LIG_PX_ARA = 40, 41, 42, 43, 44, 45

sec(hy, 47, "F. PARAMETRES FINANCIERS, SOCIAUX ET FISCAUX", 5)
entete(hy, 48, ["Parametre", "Campagne 2024", "Campagne 2025", "", "Commentaire"])
finparam = [
    ("Taux d'impot sur les benefices", 0.25, 0.25, FMT_PCT, "Taux de droit commun, hors regime derogatoire agricole"),
    ("Taux d'interet emprunt de campagne / moyen terme", 0.14, 0.14, FMT_PCT, "Conditions du credit rural en Guinee"),
    ("Taux d'interet emprunt d'investissement", 0.00, 0.13, FMT_PCT, "Ligne d'investissement 2025, differe de 12 mois"),
    ("Taux de charges sociales patronales (CNSS)", 0.18, 0.18, FMT_PCT, "Cotisation employeur sur salaires bruts"),
    ("Cout moyen de l'homme-jour (GNF)", 30000, 32000, FMT_GNF, "Main-d'oeuvre journaliere locale, repas inclus"),
    ("Nombre d'hommes-jours de la campagne", 2800, 3200, '#,##0', "Semis, sarclages, recolte, battage, conditionnement"),
    ("Taux de change indicatif (GNF / USD)", 8600, 8700, FMT_GNF, "Pour conversion et lecture par un partenaire externe"),
]
r = 49
for lib, v24, v25, fmt, com in finparam:
    hy.cell(row=r, column=1, value=lib).font = F_NORMAL
    for i, v in enumerate([v24, v25]):
        c = hy.cell(row=r, column=2 + i, value=v)
        c.font = F_INPUT
        c.number_format = fmt
    hy.cell(row=r, column=5, value=com).font = F_NOTE
    for c in range(1, 6):
        hy.cell(row=r, column=c).border = BORDURE
    r += 1
LIG_IS, LIG_TX_MT, LIG_TX_INV, LIG_CNSS, LIG_HJ_COUT, LIG_HJ_NB, LIG_FX = 49, 50, 51, 52, 53, 54, 55

hy.cell(row=57, column=1,
        value="Sources des prix et des rendements : bareme indicatif de campagne des marches de Lola et de "
              "N'Zerekore, references techniques ANPROCA / IRAG pour la Guinee forestiere, et estimations de "
              "l'exploitant. Ces valeurs sont des hypotheses de travail a valider avant tout engagement.").font = F_NOTE
hy.merge_cells(start_row=57, start_column=1, end_row=57, end_column=5)
hy.row_dimensions[57].height = 30
hy.cell(row=57, column=1).alignment = Alignment(wrap_text=True, vertical="top")

H = "Hypotheses"


# =====================================================================
# 3. PRODUCTION ET PRODUITS
# =====================================================================
pr = wb.create_sheet("Production")
pr.sheet_view.showGridLines = False
pr.column_dimensions["A"].width = 52
for col in "BCDEFGHI":
    pr.column_dimensions[col].width = 16
pr.column_dimensions["J"].width = 44

titre_page(pr, "PRODUCTION, AFFECTATION ET CHIFFRE D'AFFAIRES", "Tonnages en tonnes, valeurs en GNF", 9)

sec(pr, 4, "1. PRODUCTION BRUTE ET PRODUCTION COMMERCIALISABLE", 10)
entete(pr, 5, ["Speculation", "Surface 2024 (ha)", "Rendt 2024 (t/ha)",
               "Prod. brute 2024 (t)", "Prod. nette 2024 (t)",
               "Surface 2025 (ha)", "Rendt 2025 (t/ha)",
               "Prod. brute 2025 (t)", "Prod. nette 2025 (t)"])
cultures = ["Riz pluvial NERICA", "Riz de bas-fond amenage", "Mais grain",
            "Manioc (racines fraiches)", "Arachide (gousses en coque)"]
r = 6
for i, lib in enumerate(cultures):
    hr = LIG_ASSOL_DEB + i      # ligne assolement
    yr = LIG_RDT_DEB + i        # ligne rendement
    pr.cell(row=r, column=1, value=lib).font = F_NORMAL
    pr.cell(row=r, column=2, value=f"='{H}'!B{hr}").font = F_LIEN
    pr.cell(row=r, column=3, value=f"='{H}'!B{yr}").font = F_LIEN
    pr.cell(row=r, column=4, value=f"=B{r}*C{r}").font = F_NORMAL
    pr.cell(row=r, column=5, value=f"=D{r}*(1-'{H}'!$B${LIG_PERTES})").font = F_NORMAL
    pr.cell(row=r, column=6, value=f"='{H}'!C{hr}").font = F_LIEN
    pr.cell(row=r, column=7, value=f"='{H}'!C{yr}").font = F_LIEN
    pr.cell(row=r, column=8, value=f"=F{r}*G{r}").font = F_NORMAL
    pr.cell(row=r, column=9, value=f"=H{r}*(1-'{H}'!$C${LIG_PERTES})").font = F_NORMAL
    for c in range(2, 10):
        pr.cell(row=r, column=c).number_format = FMT_T if c not in (2, 6) else FMT_HA
    for c in range(1, 10):
        pr.cell(row=r, column=c).border = BORDURE
    r += 1
P_RIZP, P_RIZB, P_MAIS, P_MANIOC, P_ARA = 6, 7, 8, 9, 10
pr.cell(row=11, column=1, value="Total").font = F_TOTAL
for c in range(2, 10):
    cc = pr.cell(row=11, column=c, value=f"=SUM({get_column_letter(c)}6:{get_column_letter(c)}10)")
    cc.font = F_TOTAL
    cc.fill = FILL_TOTAL
    cc.number_format = FMT_T if c not in (2, 6) else FMT_HA
    cc.border = BORDURE
pr.cell(row=11, column=1).fill = FILL_TOTAL
pr.cell(row=11, column=1).border = BORDURE

sec(pr, 13, "2. AFFECTATION ET VALORISATION DU RIZ", 10)
entete(pr, 14, ["Poste", "Campagne 2024", "Campagne 2025", "", "", "", "", "", "", "Commentaire"])
riz = [
    ("Paddy net disponible (t)", f"=E{P_RIZP}+E{P_RIZB}", f"=I{P_RIZP}+I{P_RIZB}", FMT_T,
     "Somme du riz pluvial et du riz de bas-fond, apres pertes"),
    ("Paddy dirige vers l'usinage (t)", f"=B15*'{H}'!B{LIG_PART_USINE}", f"=C15*'{H}'!C{LIG_PART_USINE}", FMT_T,
     "2024 : prestation externe. 2025 : decortiqueuse propre"),
    ("Riz blanc obtenu (t)", f"=B16*'{H}'!B{LIG_USINAGE}", f"=C16*'{H}'!C{LIG_USINAGE}", FMT_T,
     "Rendement d'usinage applique au paddy usine"),
    ("Paddy vendu en l'etat (t)", "=B15-B16", "=C15-C16", FMT_T,
     "Solde vendu brut aux collecteurs"),
    ("Valeur du riz blanc (GNF)", f"=B17*'{H}'!B{LIG_PX_RIZ}*1000", f"=C17*'{H}'!C{LIG_PX_RIZ}*1000", FMT_GNF,
     "Tonnes x prix au kg x 1 000"),
    ("Valeur du paddy vendu (GNF)", f"=B18*'{H}'!B{LIG_PX_PADDY}*1000", f"=C18*'{H}'!C{LIG_PX_PADDY}*1000", FMT_GNF,
     ""),
    ("Chiffre d'affaires riz (GNF)", "=B19+B20", "=C19+C20", FMT_GNF, ""),
]
r = 15
for lib, f24, f25, fmt, com in riz:
    est_total = lib.startswith("Chiffre")
    ligne(pr, r, lib, [f24, f25], fmt,
          F_TOTAL if est_total else F_NORMAL, FILL_TOTAL if est_total else None)
    pr.cell(row=r, column=10, value=com).font = F_NOTE
    r += 1
CA_RIZ = 21

sec(pr, 23, "3. AFFECTATION ET VALORISATION DU MANIOC", 10)
entete(pr, 24, ["Poste", "Campagne 2024", "Campagne 2025", "", "", "", "", "", "", "Commentaire"])
manioc = [
    ("Racines nettes disponibles (t)", f"=E{P_MANIOC}", f"=I{P_MANIOC}", FMT_T, ""),
    ("Racines dirigees vers la transformation (t)", f"=B25*'{H}'!B{LIG_PART_GARI}", f"=C25*'{H}'!C{LIG_PART_GARI}", FMT_T,
     "2024 : prestation externe. 2025 : rapeuse-presse propre"),
    ("Gari obtenu (t)", f"=B26*'{H}'!B{LIG_GARI_TX}", f"=C26*'{H}'!C{LIG_GARI_TX}", FMT_T, ""),
    ("Racines vendues fraiches (t)", "=B25-B26", "=C25-C26", FMT_T, ""),
    ("Valeur du gari (GNF)", f"=B27*'{H}'!B{LIG_PX_GARI}*1000", f"=C27*'{H}'!C{LIG_PX_GARI}*1000", FMT_GNF, ""),
    ("Valeur des racines fraiches (GNF)", f"=B28*'{H}'!B{LIG_PX_MANIOC}*1000", f"=C28*'{H}'!C{LIG_PX_MANIOC}*1000", FMT_GNF, ""),
    ("Chiffre d'affaires manioc (GNF)", "=B29+B30", "=C29+C30", FMT_GNF, ""),
]
r = 25
for lib, f24, f25, fmt, com in manioc:
    est_total = lib.startswith("Chiffre")
    ligne(pr, r, lib, [f24, f25], fmt,
          F_TOTAL if est_total else F_NORMAL, FILL_TOTAL if est_total else None)
    pr.cell(row=r, column=10, value=com).font = F_NOTE
    r += 1
CA_MANIOC = 31

sec(pr, 33, "4. VALORISATION DU MAIS ET DE L'ARACHIDE", 10)
entete(pr, 34, ["Poste", "Campagne 2024", "Campagne 2025", "", "", "", "", "", "", "Commentaire"])
ligne(pr, 35, "Mais grain vendu (t)", [f"=E{P_MAIS}", f"=I{P_MAIS}"], FMT_T)
ligne(pr, 36, "Chiffre d'affaires mais (GNF)",
      [f"=B35*'{H}'!B{LIG_PX_MAIS}*1000", f"=C35*'{H}'!C{LIG_PX_MAIS}*1000"], FMT_GNF, F_TOTAL, FILL_TOTAL)
ligne(pr, 37, "Arachide en coque vendue (t)", [f"=E{P_ARA}", f"=I{P_ARA}"], FMT_T)
ligne(pr, 38, "Chiffre d'affaires arachide (GNF)",
      [f"=B37*'{H}'!B{LIG_PX_ARA}*1000", f"=C37*'{H}'!C{LIG_PX_ARA}*1000"], FMT_GNF, F_TOTAL, FILL_TOTAL)
CA_MAIS, CA_ARA = 36, 38

sec(pr, 40, "5. RECAPITULATIF DES PRODUITS D'EXPLOITATION (GNF)", 10)
entete(pr, 41, ["Compte SYSCOHADA / Nature du produit", "Campagne 2024", "Campagne 2025", "Variation",
                "", "", "", "", "", "Commentaire"])
recap = [
    ("701 - Ventes de riz (paddy et riz blanc)", f"=B{CA_RIZ}", f"=C{CA_RIZ}", None),
    ("701 - Ventes de mais grain", f"=B{CA_MAIS}", f"=C{CA_MAIS}", None),
    ("701 - Ventes de manioc et de gari", f"=B{CA_MANIOC}", f"=C{CA_MANIOC}", None),
    ("701 - Ventes d'arachide en coque", f"=B{CA_ARA}", f"=C{CA_ARA}", None),
    ("702 - Ventes de sous-produits (son, brisures, fanes d'arachide)", 14000000, 28000000, "input"),
    ("706 - Prestations de services (decorticage et battage a facon pour tiers)", 0, 42000000, "input"),
]
r = 42
for lib, v24, v25, typ in recap:
    ligne(pr, r, lib, [v24, v25], FMT_GNF, F_INPUT if typ == "input" else F_NORMAL)
    pr.cell(row=r, column=4, value=f'=IF(B{r}=0,"n.a.",C{r}/B{r}-1)').number_format = FMT_PCT
    pr.cell(row=r, column=4).font = F_NORMAL
    pr.cell(row=r, column=4).border = BORDURE
    r += 1
pr.cell(row=46, column=10, value="Valorisation des issues de decorticage en aliment du betail").font = F_NOTE
pr.cell(row=47, column=10, value="L'unite post-recolte acquise en 2025 travaille aussi pour les producteurs voisins").font = F_NOTE
ligne(pr, 48, "TOTAL DES PRODUITS D'EXPLOITATION", ["=SUM(B42:B47)", "=SUM(C42:C47)"],
      FMT_GNF, F_TOTAL, FILL_TOTAL)
pr.cell(row=48, column=4, value="=C48/B48-1").number_format = FMT_PCT
pr.cell(row=48, column=4).font = F_TOTAL
pr.cell(row=48, column=4).fill = FILL_TOTAL
pr.cell(row=48, column=4).border = BORDURE
LIG_TOTAL_PRODUITS = 48
PRD = "Production"

# lignes du detail riz reutilisees par les charges de prestation
LIG_PADDY_USINE = 16
LIG_MANIOC_TRANSF = 26


# =====================================================================
# 4. IMMOBILISATIONS
# =====================================================================
im = wb.create_sheet("Immobilisations")
im.sheet_view.showGridLines = False
im.column_dimensions["A"].width = 58
for col in "BCDEFG":
    im.column_dimensions[col].width = 17
im.column_dimensions["H"].width = 40

titre_page(im, "PLAN D'INVESTISSEMENT ET TABLEAU D'AMORTISSEMENT",
           "Amortissement lineaire, calcul prorata temporis a compter de la mise en service", 7)

sec(im, 4, "1. INVESTISSEMENTS DE LA CAMPAGNE 2024", 8)
entete(im, 5, ["Immobilisation", "Valeur d'acquisition (GNF)", "Duree (annees)",
               "Mois d'amort. 2024", "Dotation 2024 (GNF)", "Dotation 2025 (GNF)",
               "Amort. cumule au 31/12/2025"])
inv24 = [
    ("Amenagement et mise en valeur des terres (defriche, dessouchage, pistes, pare-feu)", 145000000, 10, 9,
     "Mise en service avril 2024"),
    ("Hangar de stockage 150 t et aire de sechage betonnee", 120000000, 20, 5, "Mise en service aout 2024"),
    ("Motoculteur et petit materiel agricole", 78000000, 5, 10, "Mise en service mars 2024"),
    ("Motopompe et equipement d'irrigation d'appoint", 25000000, 5, 8, "Mise en service mai 2024"),
    ("Tricycle de transport et motocyclettes de service", 62000000, 5, 10, "Mise en service mars 2024"),
    ("Mobilier, informatique et installations de bureau", 18000000, 5, 10, "Mise en service mars 2024"),
]
r = 6
for lib, val, duree, mois, com in inv24:
    im.cell(row=r, column=1, value=lib).font = F_NORMAL
    im.cell(row=r, column=2, value=val).font = F_INPUT
    im.cell(row=r, column=3, value=duree).font = F_INPUT
    im.cell(row=r, column=4, value=mois).font = F_INPUT
    im.cell(row=r, column=5, value=f"=B{r}/C{r}*D{r}/12").font = F_NORMAL
    im.cell(row=r, column=6, value=f"=B{r}/C{r}").font = F_NORMAL
    im.cell(row=r, column=7, value=f"=E{r}+F{r}").font = F_NORMAL
    for c in (2, 5, 6, 7):
        im.cell(row=r, column=c).number_format = FMT_GNF
    im.cell(row=r, column=8, value=com).font = F_NOTE
    for c in range(1, 8):
        im.cell(row=r, column=c).border = BORDURE
    r += 1
ligne(im, 12, "Sous-total investissements 2024",
      ["=SUM(B6:B11)", None, None, "=SUM(E6:E11)", "=SUM(F6:F11)", "=SUM(G6:G11)"],
      FMT_GNF, F_TOTAL, FILL_TOTAL)
LIG_INV24_TOTAL, LIG_DOT24_A, LIG_DOT25_A = 12, 12, 12

sec(im, 14, "2. INVESTISSEMENTS DE LA CAMPAGNE 2025", 8)
entete(im, 15, ["Immobilisation", "Valeur d'acquisition (GNF)", "Duree (annees)",
                "Mois d'amort. 2025", "Dotation 2024 (GNF)", "Dotation 2025 (GNF)",
                "Amort. cumule au 31/12/2025"])
inv25 = [
    ("Tracteur 60 CV et train d'outils (charrue, pulveriseur, remorque)", 385000000, 8, 11,
     "Mise en service fevrier 2025 - fin du recours aux prestataires"),
    ("Unite post-recolte : decortiqueuse 1 t/h, batteuse, rapeuse-presse a gari", 178000000, 7, 10,
     "Mise en service mars 2025 - internalisation de la marge d'usinage"),
    ("Extension du hangar et magasin d'intrants securise", 95000000, 20, 7, "Mise en service juin 2025"),
    ("Amenagement de 4 ha de bas-fond (planage, diguettes, canaux)", 68000000, 10, 9,
     "Mise en service avril 2025 - riziculture a maitrise d'eau"),
    ("Materiel agricole complementaire et outillage d'atelier", 32000000, 5, 8, "Mise en service mai 2025"),
]
r = 16
for lib, val, duree, mois, com in inv25:
    im.cell(row=r, column=1, value=lib).font = F_NORMAL
    im.cell(row=r, column=2, value=val).font = F_INPUT
    im.cell(row=r, column=3, value=duree).font = F_INPUT
    im.cell(row=r, column=4, value=mois).font = F_INPUT
    im.cell(row=r, column=5, value=0).font = F_NORMAL
    im.cell(row=r, column=6, value=f"=B{r}/C{r}*D{r}/12").font = F_NORMAL
    im.cell(row=r, column=7, value=f"=E{r}+F{r}").font = F_NORMAL
    for c in (2, 5, 6, 7):
        im.cell(row=r, column=c).number_format = FMT_GNF
    im.cell(row=r, column=8, value=com).font = F_NOTE
    for c in range(1, 8):
        im.cell(row=r, column=c).border = BORDURE
    r += 1
ligne(im, 21, "Sous-total investissements 2025",
      ["=SUM(B16:B20)", None, None, "=SUM(E16:E20)", "=SUM(F16:F20)", "=SUM(G16:G20)"],
      FMT_GNF, F_TOTAL, FILL_TOTAL)
LIG_INV25_TOTAL = 21

sec(im, 23, "3. SYNTHESE DES IMMOBILISATIONS", 8)
entete(im, 24, ["Poste", "Au 31/12/2024", "Au 31/12/2025", "", "", "", ""])
ligne(im, 25, "Immobilisations brutes", ["=B12", "=B12+B21"], FMT_GNF)
ligne(im, 26, "Dotation aux amortissements de l'exercice", ["=E12+E21", "=F12+F21"], FMT_GNF)
ligne(im, 27, "Amortissements cumules", ["=B26", "=B26+C26"], FMT_GNF)
ligne(im, 28, "VALEUR NETTE COMPTABLE", ["=B25-B27", "=C25-C27"], FMT_GNF, F_TOTAL, FILL_TOTAL)
LIG_IMMO_BRUTES, LIG_DOTATION, LIG_AMORT_CUM, LIG_VNC = 25, 26, 27, 28

im.cell(row=30, column=1,
        value="Le financement de ces investissements figure a l'onglet \"Flux & Financement\". Aucune "
              "immobilisation n'est cedee sur la periode : il n'y a donc ni plus-value ni moins-value de "
              "cession a constater.").font = F_NOTE
im.merge_cells(start_row=30, start_column=1, end_row=30, end_column=7)
IMM = "Immobilisations"


# =====================================================================
# 5. CHARGES
# =====================================================================
ch = wb.create_sheet("Charges")
ch.sheet_view.showGridLines = False
ch.column_dimensions["A"].width = 62
for col in "BCD":
    ch.column_dimensions[col].width = 19
ch.column_dimensions["E"].width = 48

titre_page(ch, "CHARGES D'EXPLOITATION DETAILLEES",
           "Classement par compte du plan comptable SYSCOHADA revise - montants en GNF", 5)

r = 4


def bloc_charges(titre, postes, r):
    """postes : (libelle, val2024, val2025, type, commentaire)."""
    sec(ch, r, titre, 5)
    r += 1
    entete(ch, r, ["Poste de charge", "Campagne 2024", "Campagne 2025", "Variation", "Commentaire"])
    r += 1
    debut = r
    for lib, v24, v25, typ, com in postes:
        ch.cell(row=r, column=1, value=lib).font = F_NORMAL
        for i, v in enumerate([v24, v25]):
            c = ch.cell(row=r, column=2 + i, value=v)
            c.number_format = FMT_GNF
            c.font = F_INPUT if typ == "input" else (F_LIEN if typ == "lien" else F_NORMAL)
        ch.cell(row=r, column=4, value=f'=IF(B{r}=0,"n.a.",C{r}/B{r}-1)').number_format = FMT_PCT
        ch.cell(row=r, column=4).font = F_NORMAL
        ch.cell(row=r, column=5, value=com).font = F_NOTE
        for c in range(1, 6):
            ch.cell(row=r, column=c).border = BORDURE
        r += 1
    fin = r - 1
    return debut, fin, r


# --- 60 Achats
achats = [
    ("601 - Semences et boutures", 53800000, 61460000, "input",
     "2024 : riz 1 760 kg, mais 350 kg, arachide 700 kg, boutures 12 ha. 2025 : semence certifiee R1"),
    ("602 - Engrais mineraux (NPK 15-15-15 et uree)", 106200000, 131200000, "input",
     "Riz 200 u. NPK + 100 u. uree/ha ; fumure raisonnee sur analyse de sol en 2025"),
    ("602 - Amendements et matiere organique (dolomie, fumure organique)", 0, 18000000, "input",
     "Correction de l'acidite des sols ferrallitiques, engage a partir de 2025"),
    ("602 - Produits phytosanitaires (herbicides, insecticides, traitement de semence)",
     f"='{H}'!B{LIG_EMBLAVE}*420000", f"='{H}'!C{LIG_EMBLAVE}*450000", "calc",
     "Cout unitaire par hectare emblave : 420 000 GNF en 2024, 450 000 GNF en 2025"),
    ("604 - Emballages (sacs de 50 kg, baches, ficelles, palettes)", 21000000, 32000000, "input",
     "Volume conditionne en forte hausse avec l'usinage du riz et le gari"),
    ("605 - Carburants et lubrifiants", 34000000, 96000000, "input",
     "2025 : tracteur, unite post-recolte et groupe electrogene en fonctionnement propre"),
    ("605 - Pieces de rechange, petit outillage et consommables d'atelier", 16000000, 38000000, "input",
     "Entretien du parc mecanique, en hausse avec l'acquisition du tracteur"),
]
d, f, r = bloc_charges("60 - ACHATS ET FOURNITURES CONSOMMES", achats, r)
ligne(ch, r, "Total des achats consommes", [f"=SUM(B{d}:B{f})", f"=SUM(C{d}:C{f})"],
      FMT_GNF, F_TOTAL, FILL_TOTAL)
LIG_TOT_ACHATS = r
r += 2

# --- 61/62 Services exterieurs
services = [
    ("61 - Prestations de mecanisation (labour, pulverisage, billonnage)", 48000000, 14000000, "input",
     "2024 : 40 ha travailles par prestataire. 2025 : tracteur propre, prestation residuelle"),
    ("61 - Prestation de decorticage du paddy a facon",
     f"='{PRD}'!B{LIG_PADDY_USINE}*700000", 0, "calc",
     "700 000 GNF par tonne de paddy usinee ; nul en 2025, decortiqueuse propre"),
    ("61 - Prestation de transformation du manioc en gari a facon",
     f"='{PRD}'!B{LIG_MANIOC_TRANSF}*350000", 0, "calc",
     "350 000 GNF par tonne de racines ; nul en 2025, rapeuse-presse propre"),
    ("61 - Transport, manutention et evacuation vers les marches", 46000000, 58000000, "input",
     "Gueasso vers Lola et N'Zerekore, pistes rurales, tonnage evacue en hausse"),
    ("62 - Locations et charges locatives diverses", 8000000, 6000000, "input", ""),
    ("62 - Primes d'assurance (recolte, materiel, responsabilite civile)", 6000000, 18000000, "input",
     "2025 : assurance du tracteur et de l'unite post-recolte, assurance recolte indicielle"),
    ("62 - Honoraires (expertise comptable, notaire, conseil juridique et technique)", 10000000, 16000000, "input",
     "2025 : frais de constitution de la SARL et tenue comptable reguliere"),
    ("62 - Telecommunications, energie et fournitures administratives", 12000000, 22000000, "input", ""),
    ("62 - Formation, vulgarisation et encadrement technique", 0, 12000000, "input",
     "Appui-conseil agronomique, itineraires techniques, formation des ouvriers"),
]
d, f, r = bloc_charges("61 / 62 - TRANSPORTS ET SERVICES EXTERIEURS", services, r)
ligne(ch, r, "Total des services exterieurs", [f"=SUM(B{d}:B{f})", f"=SUM(C{d}:C{f})"],
      FMT_GNF, F_TOTAL, FILL_TOTAL)
LIG_TOT_SERVICES = r
r += 2

# --- 64 Impots et taxes
impots = [
    ("64 - Contribution des patentes et licences", 4500000, 12000000, "input",
     "2024 : regime de l'entreprise individuelle. 2025 : patente de la SARL"),
    ("64 - Taxes communales, redevances foncieres et coutumieres", 9500000, 10000000, "input",
     "Redevances a la commune rurale de Gueasso et aux autorites coutumieres"),
]
d, f, r = bloc_charges("64 - IMPOTS ET TAXES", impots, r)
ligne(ch, r, "Total des impots et taxes", [f"=SUM(B{d}:B{f})", f"=SUM(C{d}:C{f})"],
      FMT_GNF, F_TOTAL, FILL_TOTAL)
LIG_TOT_IMPOTS = r
r += 2

# --- 66 Personnel
personnel = [
    ("661 - Direction de l'exploitation (1 poste)", 36000000, 43200000, "input",
     "2024 : 3 000 000 GNF x 12 mois. 2025 : 3 600 000 GNF x 12 mois"),
    ("661 - Techniciens agricoles (2 en 2024, 3 en 2025)", 40800000, 68400000, "input",
     "1 700 000 GNF x 12 x 2 puis 1 900 000 GNF x 12 x 3"),
    ("661 - Responsable transformation et commercialisation (1 poste en 2025)", 0, 26400000, "input",
     "Poste cree avec l'unite post-recolte : 2 200 000 GNF x 12 mois"),
    ("661 - Ouvriers permanents (4 en 2024, 6 en 2025)", 52800000, 97200000, "input",
     "1 200 000 GNF x 11 mois x 4 puis 1 350 000 GNF x 12 mois x 6"),
    ("661 - Gardiennage (2 agents)", 24000000, 26400000, "input",
     "Surveillance du domaine, des magasins et du parc materiel"),
]
d, f, r = bloc_charges("66 - CHARGES DE PERSONNEL", personnel, r)
LIG_SAL_DEB, LIG_SAL_FIN = d, f
ligne(ch, r, "Total des salaires bruts permanents", [f"=SUM(B{d}:B{f})", f"=SUM(C{d}:C{f})"],
      FMT_GNF, F_GRAS)
LIG_SALAIRES = r
r += 1
ligne(ch, r, "664 - Charges sociales patronales (CNSS)",
      [f"=B{LIG_SALAIRES}*'{H}'!B{LIG_CNSS}", f"=C{LIG_SALAIRES}*'{H}'!C{LIG_CNSS}"], FMT_GNF)
ch.cell(row=r, column=5, value="Taux de 18 % applique aux salaires bruts permanents").font = F_NOTE
LIG_CNSS_CH = r
r += 1
ligne(ch, r, "662 - Main-d'oeuvre journaliere de campagne",
      [f"='{H}'!B{LIG_HJ_NB}*'{H}'!B{LIG_HJ_COUT}", f"='{H}'!C{LIG_HJ_NB}*'{H}'!C{LIG_HJ_COUT}"], FMT_GNF)
ch.cell(row=r, column=5,
        value="2 800 h-j a 30 000 GNF en 2024 ; 3 200 h-j a 32 000 GNF en 2025 (semis, sarclages, recolte)").font = F_NOTE
LIG_JOURNALIERS = r
r += 1
ligne(ch, r, "Total des charges de personnel",
      [f"=B{LIG_SALAIRES}+B{LIG_CNSS_CH}+B{LIG_JOURNALIERS}",
       f"=C{LIG_SALAIRES}+C{LIG_CNSS_CH}+C{LIG_JOURNALIERS}"], FMT_GNF, F_TOTAL, FILL_TOTAL)
LIG_TOT_PERSONNEL = r
r += 2

# --- 67 Charges financieres
sec(ch, r, "67 - CHARGES FINANCIERES", 5)
r += 1
entete(ch, r, ["Poste de charge", "Campagne 2024", "Campagne 2025", "Variation", "Commentaire"])
r += 1
fin_debut = r
fin_postes = [
    ("Interets sur emprunt de campagne / moyen terme 2024",
     f"=250000000*'{H}'!B{LIG_TX_MT}*10/12", f"=208000000*'{H}'!C{LIG_TX_MT}", "calc",
     "Emprunt de 250 000 000 GNF debloque en mars 2024, 5 ans, taux de 14 %"),
    ("Interets sur emprunt d'investissement 2025",
     0, f"=450000000*'{H}'!C{LIG_TX_INV}*11/12", "calc",
     "Ligne de 450 000 000 GNF debloquee en fevrier 2025, differe de 12 mois, taux de 13 %"),
]
for lib, v24, v25, typ, com in fin_postes:
    ch.cell(row=r, column=1, value=lib).font = F_NORMAL
    for i, v in enumerate([v24, v25]):
        c = ch.cell(row=r, column=2 + i, value=v)
        c.number_format = FMT_GNF
        c.font = F_NORMAL
    ch.cell(row=r, column=4, value=f'=IF(B{r}=0,"n.a.",C{r}/B{r}-1)').number_format = FMT_PCT
    ch.cell(row=r, column=4).font = F_NORMAL
    ch.cell(row=r, column=5, value=com).font = F_NOTE
    for c in range(1, 6):
        ch.cell(row=r, column=c).border = BORDURE
    r += 1
ligne(ch, r, "Total des charges financieres",
      [f"=SUM(B{fin_debut}:B{r - 1})", f"=SUM(C{fin_debut}:C{r - 1})"], FMT_GNF, F_TOTAL, FILL_TOTAL)
LIG_TOT_FINANCIER = r
r += 2

# --- 68 Dotations
sec(ch, r, "68 - DOTATIONS AUX AMORTISSEMENTS", 5)
r += 1
entete(ch, r, ["Poste de charge", "Campagne 2024", "Campagne 2025", "Variation", "Commentaire"])
r += 1
ch.cell(row=r, column=1, value="681 - Dotations aux amortissements des immobilisations").font = F_NORMAL
ch.cell(row=r, column=2, value=f"='{IMM}'!B{LIG_DOTATION}").font = F_LIEN
ch.cell(row=r, column=3, value=f"='{IMM}'!C{LIG_DOTATION}").font = F_LIEN
for c in (2, 3):
    ch.cell(row=r, column=c).number_format = FMT_GNF
ch.cell(row=r, column=4, value=f"=C{r}/B{r}-1").number_format = FMT_PCT
ch.cell(row=r, column=5, value="Report de l'onglet Immobilisations, mode lineaire prorata temporis").font = F_NOTE
for c in range(1, 6):
    ch.cell(row=r, column=c).border = BORDURE
LIG_TOT_DOTATIONS = r
r += 2

sec(ch, r, "RECAPITULATIF GENERAL DES CHARGES", 5)
r += 1
entete(ch, r, ["Rubrique", "Campagne 2024", "Campagne 2025", "Variation", "Poids dans le total 2025"])
r += 1
recap_debut = r
for lib, src in [
    ("Achats et fournitures consommes", LIG_TOT_ACHATS),
    ("Transports et services exterieurs", LIG_TOT_SERVICES),
    ("Impots et taxes", LIG_TOT_IMPOTS),
    ("Charges de personnel", LIG_TOT_PERSONNEL),
    ("Charges financieres", LIG_TOT_FINANCIER),
    ("Dotations aux amortissements", LIG_TOT_DOTATIONS),
]:
    ch.cell(row=r, column=1, value=lib).font = F_NORMAL
    ch.cell(row=r, column=2, value=f"=B{src}").font = F_NORMAL
    ch.cell(row=r, column=3, value=f"=C{src}").font = F_NORMAL
    for c in (2, 3):
        ch.cell(row=r, column=c).number_format = FMT_GNF
    ch.cell(row=r, column=4, value=f"=C{r}/B{r}-1").number_format = FMT_PCT
    ch.cell(row=r, column=4).font = F_NORMAL
    ch.cell(row=r, column=5, value=f"=C{r}/C${recap_debut + 6}").number_format = FMT_PCT
    ch.cell(row=r, column=5).font = F_NORMAL
    for c in range(1, 6):
        ch.cell(row=r, column=c).border = BORDURE
    r += 1
ligne(ch, r, "TOTAL GENERAL DES CHARGES",
      [f"=SUM(B{recap_debut}:B{r - 1})", f"=SUM(C{recap_debut}:C{r - 1})"], FMT_GNF, F_TOTAL, FILL_TOTAL)
ch.cell(row=r, column=4, value=f"=C{r}/B{r}-1").number_format = FMT_PCT
ch.cell(row=r, column=4).font = F_TOTAL
ch.cell(row=r, column=4).fill = FILL_TOTAL
ch.cell(row=r, column=4).border = BORDURE
LIG_TOTAL_CHARGES = r
CHG = "Charges"


# =====================================================================
# 6. COMPTE DE RESULTAT
# =====================================================================
cr = wb.create_sheet("Compte de resultat")
cr.sheet_view.showGridLines = False
cr.column_dimensions["A"].width = 62
for col in "BCD":
    cr.column_dimensions[col].width = 20
cr.column_dimensions["E"].width = 14

titre_page(cr, "COMPTE DE RESULTAT COMPARATIF",
           "Campagnes agricoles 2024 et 2025 - presentation SYSCOHADA - montants en GNF", 5)

sec(cr, 4, "COMPTE DE RESULTAT DES CAMPAGNES 2024 ET 2025", 5)
entete(cr, 5, ["Libelle", "Campagne 2024", "Campagne 2025", "Variation en GNF", "Variation en %"])

cr_lignes = [
    ("PRODUITS D'EXPLOITATION", None, None, "sec"),
    ("Ventes de produits agricoles et transformes",
     f"=SUM('{PRD}'!B42:B45)", f"=SUM('{PRD}'!C42:C45)", "lien"),
    ("Ventes de sous-produits", f"='{PRD}'!B46", f"='{PRD}'!C46", "lien"),
    ("Prestations de services rendues aux tiers", f"='{PRD}'!B47", f"='{PRD}'!C47", "lien"),
    ("Total des produits d'exploitation (I)",
     f"='{PRD}'!B{LIG_TOTAL_PRODUITS}", f"='{PRD}'!C{LIG_TOTAL_PRODUITS}", "total"),
    ("", None, None, "vide"),
    ("CHARGES D'EXPLOITATION", None, None, "sec"),
    ("Achats et fournitures consommes", f"='{CHG}'!B{LIG_TOT_ACHATS}", f"='{CHG}'!C{LIG_TOT_ACHATS}", "lien"),
    ("Transports et services exterieurs", f"='{CHG}'!B{LIG_TOT_SERVICES}", f"='{CHG}'!C{LIG_TOT_SERVICES}", "lien"),
    ("Impots et taxes", f"='{CHG}'!B{LIG_TOT_IMPOTS}", f"='{CHG}'!C{LIG_TOT_IMPOTS}", "lien"),
    ("Charges de personnel", f"='{CHG}'!B{LIG_TOT_PERSONNEL}", f"='{CHG}'!C{LIG_TOT_PERSONNEL}", "lien"),
    ("Dotations aux amortissements", f"='{CHG}'!B{LIG_TOT_DOTATIONS}", f"='{CHG}'!C{LIG_TOT_DOTATIONS}", "lien"),
    ("Total des charges d'exploitation (II)", "=SUM(B13:B17)", "=SUM(C13:C17)", "total"),
    ("", None, None, "vide"),
    ("RESULTAT D'EXPLOITATION (I - II)", "=B10-B18", "=C10-C18", "grandtotal"),
    ("", None, None, "vide"),
    ("Charges financieres (III)", f"='{CHG}'!B{LIG_TOT_FINANCIER}", f"='{CHG}'!C{LIG_TOT_FINANCIER}", "lien"),
    ("RESULTAT FINANCIER (- III)", "=-B22", "=-C22", "total"),
    ("", None, None, "vide"),
    ("RESULTAT AVANT IMPOT", "=B20+B23", "=C20+C23", "grandtotal"),
    (f"Impot sur les benefices", f"=MAX(0,B25*'{H}'!B{LIG_IS})", f"=MAX(0,C25*'{H}'!C{LIG_IS})", "normal"),
    ("RESULTAT NET DE LA CAMPAGNE", "=B25-B26", "=C25-C26", "grandtotal"),
    ("", None, None, "vide"),
    ("Capacite d'autofinancement (resultat net + dotations)", "=B27+B17", "=C27+C17", "total"),
]
r = 6
for lib, f24, f25, typ in cr_lignes:
    if typ == "vide":
        r += 1
        continue
    if typ == "sec":
        sec(cr, r, lib, 5)
        r += 1
        continue
    font = {"total": F_TOTAL, "grandtotal": Font(name=POLICE, size=11, bold=True, color="C00000"),
            "lien": F_LIEN}.get(typ, F_NORMAL)
    fill = FILL_TOTAL if typ in ("total", "grandtotal") else None
    cr.cell(row=r, column=1, value=lib).font = font
    cr.cell(row=r, column=1).border = BORDURE
    if fill:
        cr.cell(row=r, column=1).fill = fill
    for i, v in enumerate([f24, f25]):
        c = cr.cell(row=r, column=2 + i, value=v)
        c.number_format = FMT_GNF
        c.font = font
        c.border = BORDURE
        if fill:
            c.fill = fill
    cr.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = FMT_GNF
    cr.cell(row=r, column=5, value=f'=IF(B{r}=0,"n.a.",C{r}/B{r}-1)').number_format = FMT_PCT
    for c in (4, 5):
        cr.cell(row=r, column=c).font = font
        cr.cell(row=r, column=c).border = BORDURE
        if fill:
            cr.cell(row=r, column=c).fill = fill
    r += 1

LIG_CR_PRODUITS = 10
LIG_CR_CHARGES_EXPL = 18
LIG_CR_RESULT_EXPL = 20
LIG_CR_FINANCIER = 22
LIG_CR_RAI = 25
LIG_CR_IS = 26
LIG_CR_RN = 27
LIG_CR_CAF = 29
LIG_CR_DOTATIONS = 17
LIG_CR_ACHATS = 13
LIG_CR_SERVICES = 14
LIG_CR_IMPOTSTAXES = 15
LIG_CR_PERSONNEL = 16

cr.cell(row=31, column=1,
        value="Les deux campagnes degagent un resultat net positif. Le resultat de 2024 est etroit parce que la "
              "campagne supporte les couts de premiere installation et une part importante de sous-traitance. "
              "En 2025, l'internalisation de la mecanisation et de la transformation, la hausse des rendements "
              "et la reduction des pertes post-recolte font plus que compenser l'alourdissement des dotations "
              "et des frais financiers lies a l'investissement.").font = F_NOTE
cr.merge_cells(start_row=31, start_column=1, end_row=33, end_column=5)
cr.cell(row=31, column=1).alignment = Alignment(wrap_text=True, vertical="top")
CRS = "Compte de resultat"


# =====================================================================
# 7. FLUX & FINANCEMENT
# =====================================================================
fl = wb.create_sheet("Flux & Financement")
fl.sheet_view.showGridLines = False
fl.column_dimensions["A"].width = 62
for col in "BC":
    fl.column_dimensions[col].width = 20
fl.column_dimensions["D"].width = 52

titre_page(fl, "TABLEAU DES FLUX DE TRESORERIE ET PLAN DE FINANCEMENT",
           "Campagnes 2024 et 2025 - montants en GNF", 4)

sec(fl, 4, "1. ELEMENTS DU BESOIN EN FONDS DE ROULEMENT", 4)
entete(fl, 5, ["Poste", "Au 31/12/2024", "Au 31/12/2025", "Commentaire"])
bfr_postes = [
    ("Stocks d'intrants et d'emballages", 28000000, 45000000, "input",
     "Engrais et semences de la campagne suivante approvisionnes en fin d'annee"),
    ("Stocks de produits finis (riz, mais, gari en magasin)", 72000000, 118000000, "input",
     "Environ 10 % de la production conservee, valorisee au cout de production"),
    ("Creances clients et autres creances", 46000000, 78000000, "input",
     "Ventes a credit court aux grossistes de N'Zerekore et Lola"),
    ("Dettes fournisseurs", 42000000, 68000000, "input", "Credit fournisseur sur intrants et transport"),
    ("Dettes sociales (salaires et CNSS a payer)", 12000000, 20000000, "input", ""),
    ("Dettes fiscales (impot sur les benefices a payer)",
     f"='{CRS}'!B{LIG_CR_IS}", f"='{CRS}'!C{LIG_CR_IS}", "lien", "Impot de l'exercice restant du a la cloture"),
]
r = 6
for lib, v24, v25, typ, com in bfr_postes:
    fl.cell(row=r, column=1, value=lib).font = F_NORMAL
    for i, v in enumerate([v24, v25]):
        c = fl.cell(row=r, column=2 + i, value=v)
        c.number_format = FMT_GNF
        c.font = F_INPUT if typ == "input" else F_LIEN
    fl.cell(row=r, column=4, value=com).font = F_NOTE
    for c in range(1, 5):
        fl.cell(row=r, column=c).border = BORDURE
    r += 1
LIG_STOCK_INTRANTS, LIG_STOCK_PF, LIG_CREANCES = 6, 7, 8
LIG_DET_FOURN, LIG_DET_SOC, LIG_DET_FISC = 9, 10, 11
ligne(fl, 12, "Besoin en fonds de roulement",
      ["=B6+B7+B8-B9-B10-B11", "=C6+C7+C8-C9-C10-C11"], FMT_GNF, F_TOTAL, FILL_TOTAL)
ligne(fl, 13, "Variation du besoin en fonds de roulement", ["=B12-0", "=C12-B12"], FMT_GNF)
LIG_BFR, LIG_VAR_BFR = 12, 13

sec(fl, 15, "2. PLAN DE FINANCEMENT", 4)
entete(fl, 16, ["Ressource / Emploi", "Campagne 2024", "Campagne 2025", "Commentaire"])
fin_postes2 = [
    ("Apports des promoteurs et comptes courants d'associes", 320000000, 50000000, "input",
     "2024 : 200 000 000 d'apports + 120 000 000 de comptes courants. 2025 : 50 000 000 complementaires"),
    ("Liberation du capital social de KASI GROUP SARL", 0, 10000000, "input",
     "Capital de 10 000 000 GNF integralement libere a la constitution du 19/11/2025"),
    ("Emprunt de campagne / moyen terme (5 ans, 14 %)", 250000000, 0, "input",
     "Debloque en mars 2024 pour l'amenagement et le fonds de roulement"),
    ("Emprunt d'investissement (7 ans, 13 %, differe de 12 mois)", 0, 450000000, "input",
     "Finance le tracteur, l'unite post-recolte et l'amenagement du bas-fond"),
    ("Remboursement du principal des emprunts", 42000000, 42000000, "input",
     "Amortissement du principal de l'emprunt 2024"),
]
r = 17
for lib, v24, v25, typ, com in fin_postes2:
    fl.cell(row=r, column=1, value=lib).font = F_NORMAL
    for i, v in enumerate([v24, v25]):
        c = fl.cell(row=r, column=2 + i, value=v)
        c.number_format = FMT_GNF
        c.font = F_INPUT
    fl.cell(row=r, column=4, value=com).font = F_NOTE
    for c in range(1, 5):
        fl.cell(row=r, column=c).border = BORDURE
    r += 1
LIG_APPORTS, LIG_CAPITAL, LIG_EMP_MT, LIG_EMP_INV, LIG_REMB = 17, 18, 19, 20, 21

ligne(fl, 22, "Encours de l'emprunt moyen terme a la cloture", ["=B19-B21", "=B22-C21"], FMT_GNF, F_GRAS)
ligne(fl, 23, "Encours de l'emprunt d'investissement a la cloture", ["=B20", "=B23+C20"], FMT_GNF, F_GRAS)
LIG_ENC_MT, LIG_ENC_INV = 22, 23

sec(fl, 25, "3. TABLEAU DES FLUX DE TRESORERIE", 4)
entete(fl, 26, ["Flux", "Campagne 2024", "Campagne 2025", "Commentaire"])
flux = [
    ("Tresorerie a l'ouverture", 0, "=B38", "calc", "Exploitation demarree en 2024, tresorerie initiale nulle"),
    ("Capacite d'autofinancement", f"='{CRS}'!B{LIG_CR_CAF}", f"='{CRS}'!C{LIG_CR_CAF}", "lien",
     "Resultat net augmente des dotations aux amortissements"),
    ("Variation du besoin en fonds de roulement", f"=-B{LIG_VAR_BFR}", f"=-C{LIG_VAR_BFR}", "calc", ""),
    ("Flux net de tresorerie d'exploitation", "=B28+B29", "=C28+C29", "total", ""),
    ("Acquisitions d'immobilisations", f"=-'{IMM}'!B{LIG_INV24_TOTAL}", f"=-'{IMM}'!B{LIG_INV25_TOTAL}", "lien", ""),
    ("Flux net de tresorerie d'investissement", "=B31", "=C31", "total", ""),
    ("Apports et comptes courants d'associes", f"=B{LIG_APPORTS}+B{LIG_CAPITAL}",
     f"=C{LIG_APPORTS}+C{LIG_CAPITAL}", "calc", ""),
    ("Emprunts nouveaux", f"=B{LIG_EMP_MT}+B{LIG_EMP_INV}", f"=C{LIG_EMP_MT}+C{LIG_EMP_INV}", "calc", ""),
    ("Remboursement du principal des emprunts", f"=-B{LIG_REMB}", f"=-C{LIG_REMB}", "calc", ""),
    ("Flux net de tresorerie de financement", "=SUM(B33:B35)", "=SUM(C33:C35)", "total", ""),
    ("Variation de tresorerie de la campagne", "=B30+B32+B36", "=C30+C32+C36", "total", ""),
    ("TRESORERIE A LA CLOTURE", "=B27+B37", "=C27+C37", "grandtotal", "Disponibilites en banque et en caisse"),
]
r = 27
for lib, f24, f25, typ, com in flux:
    font = {"total": F_TOTAL, "grandtotal": Font(name=POLICE, size=11, bold=True, color="C00000"),
            "lien": F_LIEN}.get(typ, F_NORMAL)
    fill = FILL_TOTAL if typ in ("total", "grandtotal") else None
    fl.cell(row=r, column=1, value=lib).font = font
    fl.cell(row=r, column=1).border = BORDURE
    if fill:
        fl.cell(row=r, column=1).fill = fill
    for i, v in enumerate([f24, f25]):
        c = fl.cell(row=r, column=2 + i, value=v)
        c.number_format = FMT_GNF
        c.font = font
        c.border = BORDURE
        if fill:
            c.fill = fill
    fl.cell(row=r, column=4, value=com).font = F_NOTE
    r += 1
LIG_TRESO_CLOTURE = 38
FLX = "Flux & Financement"


# =====================================================================
# 8. BILAN
# =====================================================================
bi = wb.create_sheet("Bilan")
bi.sheet_view.showGridLines = False
bi.column_dimensions["A"].width = 62
for col in "BCD":
    bi.column_dimensions[col].width = 20
bi.column_dimensions["E"].width = 40

titre_page(bi, "BILAN COMPARATIF AU 31 DECEMBRE 2024 ET AU 31 DECEMBRE 2025",
           "Presentation SYSCOHADA - montants en GNF", 5)

sec(bi, 4, "ACTIF", 5)
entete(bi, 5, ["Poste d'actif", "Au 31/12/2024", "Au 31/12/2025", "Variation", "Commentaire"])
actif = [
    ("ACTIF IMMOBILISE", None, None, "sec2", ""),
    ("Immobilisations brutes", f"='{IMM}'!B{LIG_IMMO_BRUTES}", f"='{IMM}'!C{LIG_IMMO_BRUTES}", "lien", ""),
    ("Amortissements cumules (a deduire)", f"=-'{IMM}'!B{LIG_AMORT_CUM}", f"=-'{IMM}'!C{LIG_AMORT_CUM}", "lien", ""),
    ("Total de l'actif immobilise net", "=B7+B8", "=C7+C8", "total", ""),
    ("ACTIF CIRCULANT", None, None, "sec2", ""),
    ("Stocks d'intrants et d'emballages", f"='{FLX}'!B{LIG_STOCK_INTRANTS}", f"='{FLX}'!C{LIG_STOCK_INTRANTS}", "lien", ""),
    ("Stocks de produits finis", f"='{FLX}'!B{LIG_STOCK_PF}", f"='{FLX}'!C{LIG_STOCK_PF}", "lien", ""),
    ("Creances clients et autres creances", f"='{FLX}'!B{LIG_CREANCES}", f"='{FLX}'!C{LIG_CREANCES}", "lien", ""),
    ("Total de l'actif circulant", "=SUM(B11:B13)", "=SUM(C11:C13)", "total", ""),
    ("TRESORERIE - ACTIF", None, None, "sec2", ""),
    ("Banques, caisse et disponibilites", f"='{FLX}'!B{LIG_TRESO_CLOTURE}", f"='{FLX}'!C{LIG_TRESO_CLOTURE}", "lien", ""),
    ("TOTAL DE L'ACTIF", "=B9+B14+B16", "=C9+C14+C16", "grandtotal", ""),
]
r = 6
for lib, f24, f25, typ, com in actif:
    if typ == "sec2":
        bi.cell(row=r, column=1, value=lib).font = Font(name=POLICE, size=10, bold=True, color="1F3864")
        for c in range(1, 6):
            bi.cell(row=r, column=c).fill = PatternFill("solid", fgColor="DEEBF7")
            bi.cell(row=r, column=c).border = BORDURE
        r += 1
        continue
    font = {"total": F_TOTAL, "grandtotal": Font(name=POLICE, size=11, bold=True, color="C00000"),
            "lien": F_LIEN}.get(typ, F_NORMAL)
    fill = FILL_TOTAL if typ in ("total", "grandtotal") else None
    bi.cell(row=r, column=1, value=lib).font = font
    bi.cell(row=r, column=1).border = BORDURE
    if fill:
        bi.cell(row=r, column=1).fill = fill
    for i, v in enumerate([f24, f25]):
        c = bi.cell(row=r, column=2 + i, value=v)
        c.number_format = FMT_GNF
        c.font = font
        c.border = BORDURE
        if fill:
            c.fill = fill
    bi.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = FMT_GNF
    bi.cell(row=r, column=4).font = font
    bi.cell(row=r, column=4).border = BORDURE
    if fill:
        bi.cell(row=r, column=4).fill = fill
    bi.cell(row=r, column=5, value=com).font = F_NOTE
    r += 1
LIG_TOTAL_ACTIF = 17

sec(bi, 20, "PASSIF", 5)
entete(bi, 21, ["Poste de passif", "Au 31/12/2024", "Au 31/12/2025", "Variation", "Commentaire"])
passif = [
    ("CAPITAUX PROPRES ET RESSOURCES ASSIMILEES", None, None, "sec2", ""),
    ("Capital social libere", 0, f"='{FLX}'!C{LIG_CAPITAL}", "lien",
     "10 000 000 GNF liberes a la constitution du 19/11/2025"),
    ("Apports des promoteurs (exploitation en nom propre)", 200000000, 0, "input",
     "Reclasses en comptes courants d'associes lors de l'apport a la SARL en 2025"),
    ("Report a nouveau (resultat de la campagne 2024)", 0, f"='{CRS}'!B{LIG_CR_RN}", "lien",
     "Resultat 2024 non distribue, affecte en report a nouveau"),
    ("Resultat net de la campagne", f"='{CRS}'!B{LIG_CR_RN}", f"='{CRS}'!C{LIG_CR_RN}", "lien", ""),
    ("Total des capitaux propres", "=SUM(B23:B26)", "=SUM(C23:C26)", "total", ""),
    ("Comptes courants d'associes bloques", 120000000, 370000000, "input",
     "2025 : 200 000 000 reclasses + 120 000 000 anterieurs + 50 000 000 nouveaux"),
    ("Capitaux propres elargis (y compris comptes courants bloques)", "=B27+B28", "=C27+C28", "total",
     "Agregat retenu par les bailleurs pour apprecier l'autonomie financiere"),
    ("DETTES FINANCIERES", None, None, "sec2", ""),
    ("Emprunt moyen terme 2024 (encours)", f"='{FLX}'!B{LIG_ENC_MT}", f"='{FLX}'!C{LIG_ENC_MT}", "lien", ""),
    ("Emprunt d'investissement 2025 (encours)", f"='{FLX}'!B{LIG_ENC_INV}", f"='{FLX}'!C{LIG_ENC_INV}", "lien", ""),
    ("Total des dettes financieres", "=B31+B32", "=C31+C32", "total", ""),
    ("PASSIF CIRCULANT", None, None, "sec2", ""),
    ("Dettes fournisseurs", f"='{FLX}'!B{LIG_DET_FOURN}", f"='{FLX}'!C{LIG_DET_FOURN}", "lien", ""),
    ("Dettes sociales", f"='{FLX}'!B{LIG_DET_SOC}", f"='{FLX}'!C{LIG_DET_SOC}", "lien", ""),
    ("Dettes fiscales (impot sur les benefices)", f"='{FLX}'!B{LIG_DET_FISC}", f"='{FLX}'!C{LIG_DET_FISC}", "lien", ""),
    ("Total du passif circulant", "=SUM(B35:B37)", "=SUM(C35:C37)", "total", ""),
    ("TOTAL DU PASSIF", "=B29+B33+B38", "=C29+C33+C38", "grandtotal", ""),
]
r = 22
for lib, f24, f25, typ, com in passif:
    if typ == "sec2":
        bi.cell(row=r, column=1, value=lib).font = Font(name=POLICE, size=10, bold=True, color="1F3864")
        for c in range(1, 6):
            bi.cell(row=r, column=c).fill = PatternFill("solid", fgColor="DEEBF7")
            bi.cell(row=r, column=c).border = BORDURE
        r += 1
        continue
    font = {"total": F_TOTAL, "grandtotal": Font(name=POLICE, size=11, bold=True, color="C00000"),
            "lien": F_LIEN, "input": F_INPUT}.get(typ, F_NORMAL)
    fill = FILL_TOTAL if typ in ("total", "grandtotal") else None
    bi.cell(row=r, column=1, value=lib).font = font
    bi.cell(row=r, column=1).border = BORDURE
    if fill:
        bi.cell(row=r, column=1).fill = fill
    for i, v in enumerate([f24, f25]):
        c = bi.cell(row=r, column=2 + i, value=v)
        c.number_format = FMT_GNF
        c.font = font
        c.border = BORDURE
        if fill:
            c.fill = fill
    bi.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = FMT_GNF
    bi.cell(row=r, column=4).font = font
    bi.cell(row=r, column=4).border = BORDURE
    if fill:
        bi.cell(row=r, column=4).fill = fill
    bi.cell(row=r, column=5, value=com).font = F_NOTE
    r += 1
LIG_CP, LIG_CC, LIG_CP_ELARGI = 27, 28, 29
LIG_DETTES_FIN, LIG_PASSIF_CIRC, LIG_TOTAL_PASSIF = 33, 38, 39

sec(bi, 42, "CONTROLE D'EQUILIBRE DU BILAN", 5)
ligne(bi, 43, "Total de l'actif", [f"=B{LIG_TOTAL_ACTIF}", f"=C{LIG_TOTAL_ACTIF}"], FMT_GNF, F_GRAS)
ligne(bi, 44, "Total du passif", [f"=B{LIG_TOTAL_PASSIF}", f"=C{LIG_TOTAL_PASSIF}"], FMT_GNF, F_GRAS)
ligne(bi, 45, "Ecart actif - passif (doit etre nul)", ["=B43-B44", "=C43-C44"], FMT_GNF, F_TOTAL, FILL_TOTAL)
bi.cell(row=46, column=1, value='=IF(AND(ABS(B45)<1,ABS(C45)<1),"BILAN EQUILIBRE SUR LES DEUX CAMPAGNES",'
                                '"DESEQUILIBRE - VERIFIER LES SAISIES")')
bi.cell(row=46, column=1).font = Font(name=POLICE, size=11, bold=True, color="006100")
bi.cell(row=46, column=1).fill = PatternFill("solid", fgColor="C6EFCE")

sec(bi, 48, "EQUILIBRE FINANCIER", 5)
entete(bi, 49, ["Indicateur", "Au 31/12/2024", "Au 31/12/2025", "", "Lecture"])
ligne(bi, 50, "Ressources stables (capitaux propres elargis + dettes financieres)",
      [f"=B{LIG_CP_ELARGI}+B{LIG_DETTES_FIN}", f"=C{LIG_CP_ELARGI}+C{LIG_DETTES_FIN}"], FMT_GNF)
ligne(bi, 51, "Actif immobilise net", ["=B9", "=C9"], FMT_GNF)
ligne(bi, 52, "Fonds de roulement net global", ["=B50-B51", "=C50-C51"], FMT_GNF, F_TOTAL, FILL_TOTAL)
ligne(bi, 53, "Besoin en fonds de roulement", [f"='{FLX}'!B{LIG_BFR}", f"='{FLX}'!C{LIG_BFR}"], FMT_GNF)
ligne(bi, 54, "Tresorerie nette (fonds de roulement - besoin en fonds de roulement)",
      ["=B52-B53", "=C52-C53"], FMT_GNF, F_TOTAL, FILL_TOTAL)
bi.cell(row=52, column=5, value="Positif : les emplois durables sont finances par des ressources durables").font = F_NOTE
bi.cell(row=54, column=5, value="Doit egaler la tresorerie inscrite a l'actif").font = F_NOTE
BLN = "Bilan"


# =====================================================================
# 9. SIG & RATIOS
# =====================================================================
sg = wb.create_sheet("SIG & Ratios")
sg.sheet_view.showGridLines = False
sg.column_dimensions["A"].width = 62
for col in "BCD":
    sg.column_dimensions[col].width = 20
sg.column_dimensions["E"].width = 46

titre_page(sg, "SOLDES INTERMEDIAIRES DE GESTION, RATIOS ET SEUIL DE RENTABILITE",
           "Campagnes 2024 et 2025", 5)

sec(sg, 4, "1. SOLDES INTERMEDIAIRES DE GESTION (GNF)", 5)
entete(sg, 5, ["Solde", "Campagne 2024", "Campagne 2025", "Variation en %", "Definition"])
sig = [
    ("Chiffre d'affaires et produits d'exploitation",
     f"='{CRS}'!B{LIG_CR_PRODUITS}", f"='{CRS}'!C{LIG_CR_PRODUITS}",
     "Total des ventes, sous-produits et prestations"),
    ("Consommations intermediaires",
     f"='{CRS}'!B{LIG_CR_ACHATS}+'{CRS}'!B{LIG_CR_SERVICES}",
     f"='{CRS}'!C{LIG_CR_ACHATS}+'{CRS}'!C{LIG_CR_SERVICES}",
     "Achats consommes et services exterieurs"),
    ("VALEUR AJOUTEE", "=B6-B7", "=C6-C7", "Richesse creee par l'exploitation"),
    ("Impots et taxes", f"='{CRS}'!B{LIG_CR_IMPOTSTAXES}", f"='{CRS}'!C{LIG_CR_IMPOTSTAXES}", ""),
    ("Charges de personnel", f"='{CRS}'!B{LIG_CR_PERSONNEL}", f"='{CRS}'!C{LIG_CR_PERSONNEL}", ""),
    ("EXCEDENT BRUT D'EXPLOITATION", "=B8-B9-B10", "=C8-C9-C10",
     "Performance economique avant amortissements et financement"),
    ("Dotations aux amortissements", f"='{CRS}'!B{LIG_CR_DOTATIONS}", f"='{CRS}'!C{LIG_CR_DOTATIONS}", ""),
    ("RESULTAT D'EXPLOITATION", "=B11-B12", "=C11-C12", ""),
    ("Charges financieres", f"='{CRS}'!B{LIG_CR_FINANCIER}", f"='{CRS}'!C{LIG_CR_FINANCIER}", ""),
    ("RESULTAT AVANT IMPOT", "=B13-B14", "=C13-C14", ""),
    ("Impot sur les benefices", f"='{CRS}'!B{LIG_CR_IS}", f"='{CRS}'!C{LIG_CR_IS}", ""),
    ("RESULTAT NET", "=B15-B16", "=C15-C16", ""),
    ("CAPACITE D'AUTOFINANCEMENT", "=B17+B12", "=C17+C12", "Resultat net augmente des dotations"),
]
r = 6
for lib, f24, f25, com in sig:
    maj = lib.isupper()
    ligne(sg, r, lib, [f24, f25], FMT_GNF, F_TOTAL if maj else F_NORMAL, FILL_TOTAL if maj else None)
    sg.cell(row=r, column=4, value=f'=IF(B{r}=0,"n.a.",C{r}/B{r}-1)').number_format = FMT_PCT
    sg.cell(row=r, column=4).font = F_TOTAL if maj else F_NORMAL
    sg.cell(row=r, column=4).border = BORDURE
    if maj:
        sg.cell(row=r, column=4).fill = FILL_TOTAL
    sg.cell(row=r, column=5, value=com).font = F_NOTE
    r += 1
LIG_VA, LIG_EBE, LIG_RE_SIG, LIG_RN_SIG, LIG_CAF_SIG = 8, 11, 13, 17, 18

sec(sg, 25, "2. RATIOS DE RENTABILITE", 5)
entete(sg, 26, ["Ratio", "Campagne 2024", "Campagne 2025", "", "Lecture"])
ratios_rent = [
    ("Taux de valeur ajoutee (valeur ajoutee / produits)", "=B8/B6", "=C8/C6", FMT_PCT,
     "Part de la richesse conservee par l'exploitation"),
    ("Taux de marge brute d'exploitation (EBE / produits)", "=B11/B6", "=C11/C6", FMT_PCT,
     "Au-dela de 25 %, l'exploitation est structurellement saine"),
    ("Taux de marge nette (resultat net / produits)", "=B17/B6", "=C17/C6", FMT_PCT, ""),
    ("Rentabilite economique (resultat d'exploitation / total du bilan)",
     f"=B13/'{BLN}'!B{LIG_TOTAL_ACTIF}", f"=C13/'{BLN}'!C{LIG_TOTAL_ACTIF}", FMT_PCT, ""),
    ("Rentabilite des capitaux propres elargis (resultat net / capitaux propres elargis)",
     f"=B17/'{BLN}'!B{LIG_CP_ELARGI}", f"=C17/'{BLN}'!C{LIG_CP_ELARGI}", FMT_PCT, ""),
]
r = 27
for lib, f24, f25, fmt, com in ratios_rent:
    ligne(sg, r, lib, [f24, f25], fmt)
    sg.cell(row=r, column=5, value=com).font = F_NOTE
    r += 1

sec(sg, 33, "3. RATIOS DE STRUCTURE ET DE SOLVABILITE", 5)
entete(sg, 34, ["Ratio", "Au 31/12/2024", "Au 31/12/2025", "", "Lecture"])
ratios_struct = [
    ("Autonomie financiere (capitaux propres elargis / total du bilan)",
     f"='{BLN}'!B{LIG_CP_ELARGI}/'{BLN}'!B{LIG_TOTAL_ACTIF}",
     f"='{BLN}'!C{LIG_CP_ELARGI}/'{BLN}'!C{LIG_TOTAL_ACTIF}", FMT_PCT,
     "Le repli de 2025 traduit l'effet de levier de l'annee d'investissement, non une degradation"),
    ("Taux d'endettement financier (dettes financieres / capitaux propres elargis)",
     f"='{BLN}'!B{LIG_DETTES_FIN}/'{BLN}'!B{LIG_CP_ELARGI}",
     f"='{BLN}'!C{LIG_DETTES_FIN}/'{BLN}'!C{LIG_CP_ELARGI}", FMT_PCT, ""),
    ("Capacite de remboursement (dettes financieres / capacite d'autofinancement, en annees)",
     f"='{BLN}'!B{LIG_DETTES_FIN}/B18", f"='{BLN}'!C{LIG_DETTES_FIN}/C18", FMT_AN,
     "Norme bancaire : moins de 4 annees. Les deux campagnes sont tres en deca"),
    ("Couverture des frais financiers (EBE / charges financieres)",
     "=B11/B14", "=C11/C14", FMT_AN, "Norme : superieur a 3. Marge de securite confortable"),
    ("Liquidite generale (actif circulant + tresorerie / passif circulant)",
     f"=('{BLN}'!B14+'{BLN}'!B16)/'{BLN}'!B{LIG_PASSIF_CIRC}",
     f"=('{BLN}'!C14+'{BLN}'!C16)/'{BLN}'!C{LIG_PASSIF_CIRC}", FMT_AN,
     "Norme : superieur a 1,5"),
]
r = 35
for lib, f24, f25, fmt, com in ratios_struct:
    ligne(sg, r, lib, [f24, f25], fmt)
    sg.cell(row=r, column=5, value=com).font = F_NOTE
    r += 1

sec(sg, 41, "4. PERFORMANCE RAMENEE A L'HECTARE (GNF / ha emblave)", 5)
entete(sg, 42, ["Indicateur", "Campagne 2024", "Campagne 2025", "Variation en %", "Lecture"])
perha = [
    ("Superficie emblavee (ha)", f"='{H}'!B{LIG_EMBLAVE}", f"='{H}'!C{LIG_EMBLAVE}", FMT_HA, ""),
    ("Produits d'exploitation par hectare", "=B6/B43", "=C6/C43", FMT_GNF, ""),
    ("Valeur ajoutee par hectare", "=B8/B43", "=C8/C43", FMT_GNF, "Indicateur central de la productivite du foncier"),
    ("Excedent brut d'exploitation par hectare", "=B11/B43", "=C11/C43", FMT_GNF, ""),
    ("Resultat net par hectare", "=B17/B43", "=C17/C43", FMT_GNF, ""),
    ("Charges d'exploitation par hectare", "=(B7+B9+B10+B12)/B43", "=(C7+C9+C10+C12)/C43", FMT_GNF, ""),
]
r = 43
for lib, f24, f25, fmt, com in perha:
    ligne(sg, r, lib, [f24, f25], fmt)
    sg.cell(row=r, column=4, value=f'=IF(B{r}=0,"n.a.",C{r}/B{r}-1)').number_format = FMT_PCT
    sg.cell(row=r, column=4).font = F_NORMAL
    sg.cell(row=r, column=4).border = BORDURE
    sg.cell(row=r, column=5, value=com).font = F_NOTE
    r += 1

sec(sg, 50, "5. SEUIL DE RENTABILITE", 5)
entete(sg, 51, ["Element", "Campagne 2024", "Campagne 2025", "", "Methode"])
ligne(sg, 52, "Charges fixes (personnel permanent, dotations, frais financiers, impots et taxes, services fixes)",
      [f"='{CHG}'!B{LIG_SALAIRES}+'{CHG}'!B{LIG_CNSS_CH}+'{CHG}'!B{LIG_TOT_DOTATIONS}"
       f"+'{CHG}'!B{LIG_TOT_FINANCIER}+'{CHG}'!B{LIG_TOT_IMPOTS}+'{CHG}'!B21+'{CHG}'!B22+'{CHG}'!B23"
       f"+'{CHG}'!B24+'{CHG}'!B25",
       f"='{CHG}'!C{LIG_SALAIRES}+'{CHG}'!C{LIG_CNSS_CH}+'{CHG}'!C{LIG_TOT_DOTATIONS}"
       f"+'{CHG}'!C{LIG_TOT_FINANCIER}+'{CHG}'!C{LIG_TOT_IMPOTS}+'{CHG}'!C21+'{CHG}'!C22+'{CHG}'!C23"
       f"+'{CHG}'!C24+'{CHG}'!C25"], FMT_GNF)
ligne(sg, 53, "Charges variables (charges totales moins charges fixes)",
      [f"='{CHG}'!B{LIG_TOTAL_CHARGES}-B52", f"='{CHG}'!C{LIG_TOTAL_CHARGES}-C52"], FMT_GNF)
ligne(sg, 54, "Marge sur cout variable", ["=B6-B53", "=C6-C53"], FMT_GNF)
ligne(sg, 55, "Taux de marge sur cout variable", ["=B54/B6", "=C54/C6"], FMT_PCT)
ligne(sg, 56, "SEUIL DE RENTABILITE (chiffre d'affaires critique)",
      ["=B52/B55", "=C52/C55"], FMT_GNF, F_TOTAL, FILL_TOTAL)
ligne(sg, 57, "Seuil de rentabilite en pourcentage des produits", ["=B56/B6", "=C56/C6"], FMT_PCT)
ligne(sg, 58, "Marge de securite (produits - seuil de rentabilite)", ["=B6-B56", "=C6-C56"], FMT_GNF, F_TOTAL, FILL_TOTAL)
sg.cell(row=57, column=5,
        value="Plus ce pourcentage est bas, plus l'exploitation resiste a une mauvaise campagne").font = F_NOTE

sg.cell(row=60, column=1,
        value="Les charges fixes retenues comprennent les salaires permanents et les charges sociales, les "
              "dotations aux amortissements, les charges financieres, les impots et taxes, ainsi que les "
              "services exterieurs non lies au volume (locations, assurances, honoraires, telecommunications, "
              "formation). Les intrants, la main-d'oeuvre journaliere, le transport et les prestations a facon "
              "sont traites en charges variables.").font = F_NOTE
sg.merge_cells(start_row=60, start_column=1, end_row=62, end_column=5)
sg.cell(row=60, column=1).alignment = Alignment(wrap_text=True, vertical="top")
SIG = "SIG & Ratios"


# =====================================================================
# 10. COMPARATIF
# =====================================================================
cp = wb.create_sheet("Comparatif")
cp.sheet_view.showGridLines = False
cp.column_dimensions["A"].width = 56
for col in "BCDE":
    cp.column_dimensions[col].width = 20
cp.column_dimensions["F"].width = 52

titre_page(cp, "SYNTHESE COMPARATIVE DES CAMPAGNES 2024 ET 2025",
           "Deux campagnes beneficiaires, une progression construite sur la technique et la transformation", 6)

sec(cp, 4, "1. TABLEAU DE BORD DE LA PROGRESSION", 6)
entete(cp, 5, ["Indicateur", "Campagne 2024", "Campagne 2025", "Ecart en GNF",
               "Progression", "Facteur explicatif"])
tdb = [
    ("Superficie emblavee (ha)", f"='{H}'!B{LIG_EMBLAVE}", f"='{H}'!C{LIG_EMBLAVE}", FMT_HA,
     "Mise en valeur de 4 ha de bas-fond et reduction de l'emprise non cultivee"),
    ("Production totale nette (t)", f"='{PRD}'!E11", f"='{PRD}'!I11", FMT_T,
     "Rendements en hausse sur toutes les speculations, pertes divisees par deux"),
    ("Produits d'exploitation", f"='{SIG}'!B6", f"='{SIG}'!C6", FMT_GNF,
     "Effet combine des volumes, des prix et de la transformation"),
    ("Valeur ajoutee", f"='{SIG}'!B{LIG_VA}", f"='{SIG}'!C{LIG_VA}", FMT_GNF,
     "L'usinage du riz et le gari sont produits en interne au lieu d'etre sous-traites"),
    ("Excedent brut d'exploitation", f"='{SIG}'!B{LIG_EBE}", f"='{SIG}'!C{LIG_EBE}", FMT_GNF,
     "La masse salariale progresse moins vite que la valeur ajoutee"),
    ("Resultat d'exploitation", f"='{SIG}'!B{LIG_RE_SIG}", f"='{SIG}'!C{LIG_RE_SIG}", FMT_GNF, ""),
    ("RESULTAT NET DE LA CAMPAGNE", f"='{SIG}'!B{LIG_RN_SIG}", f"='{SIG}'!C{LIG_RN_SIG}", FMT_GNF,
     "Les deux campagnes sont beneficiaires ; 2025 est nettement superieure a 2024"),
    ("Capacite d'autofinancement", f"='{SIG}'!B{LIG_CAF_SIG}", f"='{SIG}'!C{LIG_CAF_SIG}", FMT_GNF,
     "Capacite de l'exploitation a financer ses investissements et a rembourser sa dette"),
    ("Total du bilan", f"='{BLN}'!B{LIG_TOTAL_ACTIF}", f"='{BLN}'!C{LIG_TOTAL_ACTIF}", FMT_GNF,
     "Doublement de l'outil de production"),
    ("Tresorerie a la cloture", f"='{FLX}'!B{LIG_TRESO_CLOTURE}", f"='{FLX}'!C{LIG_TRESO_CLOTURE}", FMT_GNF,
     "Tresorerie positive sur les deux campagnes malgre un effort d'investissement lourd"),
]
r = 6
for lib, f24, f25, fmt, com in tdb:
    maj = lib.isupper()
    ligne(cp, r, lib, [f24, f25], fmt, F_TOTAL if maj else F_NORMAL, FILL_TOTAL if maj else None)
    cp.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = fmt
    cp.cell(row=r, column=5, value=f'=IF(B{r}=0,"n.a.",C{r}/B{r}-1)').number_format = FMT_PCT
    for c in (4, 5):
        cp.cell(row=r, column=c).font = F_TOTAL if maj else F_NORMAL
        cp.cell(row=r, column=c).border = BORDURE
        if maj:
            cp.cell(row=r, column=c).fill = FILL_TOTAL
    cp.cell(row=r, column=6, value=com).font = F_NOTE
    r += 1
LIG_CP_RN = 12

sec(cp, 18, "2. DECOMPOSITION DE L'AMELIORATION DU RESULTAT", 6)
entete(cp, 19, ["Levier", "Contribution (GNF)", "", "", "", "Explication"])
leviers = [
    ("Effet volume et rendement (production nette supplementaire)",
     f"=('{PRD}'!I11-'{PRD}'!E11)/'{PRD}'!I11*('{SIG}'!C6-'{SIG}'!B6)",
     "Hausse des rendements de toutes les speculations et reduction des pertes de 10 % a 5 %"),
    ("Effet prix et transformation (valorisation unitaire supplementaire)",
     f"='{SIG}'!C6-'{SIG}'!B6-B20",
     "Usinage de 85 % du paddy contre 55 %, gari sur 45 % du manioc contre 25 %, prix en hausse"),
    ("Total de l'augmentation des produits", "=B20+B21", ""),
    ("Augmentation des charges d'exploitation et financieres",
     f"=-('{CHG}'!C{LIG_TOTAL_CHARGES}-'{CHG}'!B{LIG_TOTAL_CHARGES})",
     "Investissement, personnel renforce, dotations et frais financiers en hausse"),
    ("Augmentation de l'impot sur les benefices",
     f"=-('{CRS}'!C{LIG_CR_IS}-'{CRS}'!B{LIG_CR_IS})", ""),
    ("AMELIORATION NETTE DU RESULTAT", "=B22+B23+B24", "Ecart entre le resultat net 2025 et le resultat net 2024"),
]
r = 20
for item in leviers:
    lib, formule, com = item
    maj = lib.isupper()
    cp.cell(row=r, column=1, value=lib).font = F_TOTAL if maj else F_NORMAL
    c = cp.cell(row=r, column=2, value=formule)
    c.number_format = FMT_GNF
    c.font = F_TOTAL if maj else F_NORMAL
    for cc in range(1, 3):
        cp.cell(row=r, column=cc).border = BORDURE
        if maj:
            cp.cell(row=r, column=cc).fill = FILL_TOTAL
    cp.cell(row=r, column=6, value=com).font = F_NOTE
    r += 1
ligne(cp, 26, "Controle : resultat net 2025 moins resultat net 2024",
      [f"='{SIG}'!C{LIG_RN_SIG}-'{SIG}'!B{LIG_RN_SIG}"], FMT_GNF, F_GRAS)
cp.cell(row=27, column=1, value='=IF(ABS(B25-B26)<1,"DECOMPOSITION COHERENTE","ECART A VERIFIER")')
cp.cell(row=27, column=1).font = Font(name=POLICE, size=10, bold=True, color="006100")

sec(cp, 29, "3. CONCLUSION", 6)
conclusions = [
    "Les deux campagnes se soldent par un resultat net positif, ce qui etait l'objectif de la mise en valeur "
    "du domaine de 60 hectares de Gueasso.",
    "La campagne 2024 est beneficiaire mais etroite : elle supporte la defriche, l'amenagement initial, une "
    "mecanisation entierement sous-traitee et des pertes post-recolte de 10 %. Le seuil de rentabilite y est "
    "atteint tard dans la campagne, ce qui laisse peu de marge de securite.",
    "La campagne 2025 change d'echelle sans changer de surface. Trois leviers expliquent la progression : "
    "des rendements superieurs grace aux semences certifiees et a la fumure raisonnee, la mise en valeur de "
    "4 hectares de bas-fond rizicole a maitrise d'eau, et surtout l'internalisation de la transformation qui "
    "fait passer la part du paddy usine de 55 % a 85 % et celle du manioc transforme en gari de 25 % a 45 %.",
    "L'investissement de 2025 alourdit mecaniquement les dotations aux amortissements et les charges "
    "financieres. Cet alourdissement est plus que compense par la marge d'usinage captee en interne : la "
    "valeur ajoutee par hectare double.",
    "La structure financiere reste saine sur les deux campagnes : la tresorerie est positive a chaque cloture, "
    "le fonds de roulement couvre le besoin en fonds de roulement, et la capacite de remboursement demeure "
    "tres en deca de la norme bancaire de quatre annees.",
    "Points de vigilance a documenter avant tout engagement : la volatilite des prix du paddy et du manioc, "
    "l'irregularite pluviometrique, l'etat des pistes rurales entre Gueasso et N'Zerekore pour l'evacuation "
    "des recoltes, et la disponibilite effective de la main-d'oeuvre journaliere aux pics de travaux.",
]
r = 30
for txt in conclusions:
    c = cp.cell(row=r, column=1, value="- " + txt)
    c.font = F_NORMAL
    c.alignment = Alignment(wrap_text=True, vertical="top")
    cp.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=6)
    cp.row_dimensions[r].height = 16
    cp.row_dimensions[r + 1].height = 16
    r += 3

# ---------------------------------------------------------------- sortie
SORTIE = "/home/user/ZION/bilan-agricole-gueasso/Bilan_Comptable_Agricole_Gueasso_2024_2025.xlsx"
for feuille in wb.worksheets:
    feuille.freeze_panes = "A6"
wb["Notice"].freeze_panes = None
wb.save(SORTIE)
print("Classeur genere :", SORTIE)
